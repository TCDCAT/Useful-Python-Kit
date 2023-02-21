#Note：
#1.待导入的翻译资源需要是xlsx格式,xlsx格式参考Arabic.xlsx，需要注意
#   *第一列为strings.xml中的key，第四列为需要导入的翻译语料所在列，二三列为参考翻译
#2.使用python3运行脚本
#3.导入属于增量导入，会保留xml中的原有资源

#!/usr/bin/env python
# encoding: utf-8

import os, sys, getopt
import io
import xml.dom.minidom
import subprocess
# import xlsxwriter
import openpyxl
from xml.dom.minidom import Node



resource_excel_file_name = ""   #翻译语料excel文件路径名
resource_excel_table_name = ""  #翻译语料table名
resource_target_xml = ""        #xml文件路径名



def exportExcel2Xml():
    global resource_excel_file_name
    global resource_target_xml
    global resource_excel_table_name
    print('————开始写入xml————')
    doc = xml.dom.minidom.parse(resource_target_xml)
    strings = doc.getElementsByTagName('string')
    column_key = 1
    column_trans = 4
    # workbook = xlsxwriter.Workbook('find_strings_untranslated.xlsx', encoding ='utf-8') # 建立文件
    # worksheet = workbook.add_worksheet() # 建立sheet， 可以work.add_worksheet('employee')来指定sheet名，但中文名会报UnicodeDecodeErro的错误
    data = openpyxl.load_workbook(resource_excel_file_name)  
    table = data[resource_excel_table_name]
    nrows = len(tuple(table.rows))

    top_element = doc.documentElement
    for i in range(nrows-1):
        isTranslated = False
        key = table.cell(column=column_key, row=i+2).value
        for string in strings:
            if key == string.getAttribute('name'):
                isTranslated = True
                break

        if isTranslated == False:
            # print('unt key {0}'.format(key))
            trans_value = table.cell(column=column_trans, row=i+2).value
            # print("key:{0} hasTranslated:{1} trans_value:{2}".format(key,isTranslated,trans_value))
#            print('unt value {0}'.format(trans_value))
            # trans_value.replace('&', '&amp;')
            element = doc.createElement('string')
            element.setAttribute('name', key)
            top_element.appendChild(element)
            text = doc.createTextNode(str(trans_value))
            element.appendChild(text)
    try:
        with io.open(resource_target_xml,mode='w', encoding='UTF-8') as fh:
            # 4.writexml()第一个参数是目标文件对象，第二个参数是根节点的缩进格式，第三个参数是其他子节点的缩进格式，
            # 第四个参数制定了换行格式，第五个参数制定了xml内容的编码。
            doc.writexml(fh,indent='',addindent='\t',newl='\n',encoding='UTF-8')
            print('————写入xml成功————')
    except Exception as err:
        print('————写入xml————')
        print('错误信息：{0}'.format(err))

def readUserInput():
    global resource_excel_file_name
    global resource_target_xml
    global resource_excel_table_name
    # print("请输入xml文件绝对路径\n")
    resource_excel_file_name = input("请输入需要导入的xlsx资源文件绝对路径:")
    resource_excel_table_name = input("请输入table名:")
    # print('resource_excel_file_name {0}'.format(resource_excel_file_name))
    resource_target_xml = input("请输入导入目标xml文件绝对路径:")
    # print('resource_excel_file_name {0}'.format(resource_target_xml))
    confirm = input('请确认是否将 {0} 中的资源增量更新到 {1} 中（Y/N)？'.format(resource_excel_file_name,resource_target_xml))
    return confirm
if __name__ == '__main__':
    confirm = readUserInput()
    if confirm != "Y":
        print("Bye!")
        exit()
    else:
        exportExcel2Xml()
    # import_excel()
    # if len(sys.argv) == 1:
    #     if os.path.isfile(Axml):
    #         _check_string_res(os.path.abspath('.'))
    #     else:
    #         usage()
    # elif len(sys.argv) > 1:
    #     for path in sys.argv[1:]:
    #         if os.path.isdir(path):
    #             _check_string_res(os.path.abspath(path))
    #         else:
    #             print "### %s Not a directory, ignored." % path
    # else
    #     usage()
